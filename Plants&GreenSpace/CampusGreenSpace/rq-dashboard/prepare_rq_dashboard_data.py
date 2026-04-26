# -*- coding: utf-8 -*-
"""
Prepare lightweight JSON files for the RQ dashboard.

Inputs (provided by user):
- Plants_Shanghai_Translated.xlsx (campus flora, Shanghai campuses)
- 上海野生和逸生植物_按地区拆分.xlsx (citywide district flora, Shanghai)

Outputs (written into ./data):
- campus_summary.json
- city_district_summary.json
- overlap_family.json
- sankey_district_campus_family.json
- treemap_families.json
"""

from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl


HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"

CAMPUS_XLSX = HERE.parent / "Plants_Shanghai_Translated.xlsx"
CITY_XLSX = HERE.parent / "上海野生和逸生植物_按地区拆分.xlsx"


def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()


def _as_float(x: Any) -> float | None:
    if x is None:
        return None
    try:
        v = float(x)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None


def _is_nonnative(value: Any) -> bool | None:
    """
    Return True/False when confident, else None (unknown).
    Handles both English labels and common Chinese labels.
    """
    s = _norm(value)
    if not s:
        return None
    low = s.lower()

    # English
    if "non" in low and "native" in low:
        return True
    if low == "nonnative":
        return True
    if low == "native":
        return False

    # Chinese cues (best-effort)
    if "非" in s or "外来" in s or "归化" in s or "逸生" in s:
        return True
    if "原生" in s or "乡土" in s:
        return False

    return None


@dataclass(frozen=True)
class CampusRow:
    locality: str
    district: str
    lon: float | None
    lat: float | None
    sci_name: str
    family: str
    growth_form: str
    nativeness: str


def load_campus_rows() -> list[CampusRow]:
    wb = openpyxl.load_workbook(CAMPUS_XLSX, read_only=True, data_only=True)
    rows: list[CampusRow] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        idx = {str(h).strip(): i for i, h in enumerate(header) if h not in (None, "")}

        def col(*candidates: str) -> int:
            for c in candidates:
                if c in idx:
                    return idx[c]
            raise KeyError(f"Missing column in campus sheet {sheet}: {candidates}")

        i_locality = col("Locality")
        i_district = col("District")
        i_lon = col("Longitude")
        i_lat = col("Latitude")
        i_sci = col("ScientificName")
        i_family = col("Family")
        i_growth = col("GrowthForm")
        i_native = col("Plant_NativenessStatus")

        for r in ws.iter_rows(min_row=2, values_only=True):
            locality = _norm(r[i_locality])
            if not locality:
                continue
            rows.append(
                CampusRow(
                    locality=locality,
                    district=_norm(r[i_district]) or _norm(sheet),
                    lon=_as_float(r[i_lon]),
                    lat=_as_float(r[i_lat]),
                    sci_name=_norm(r[i_sci]),
                    family=_norm(r[i_family]) or "Unknown",
                    growth_form=_norm(r[i_growth]) or "Unknown",
                    nativeness=_norm(r[i_native]) or "Unknown",
                )
            )

    return rows


def load_city_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(CITY_XLSX, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        header = [None if h in (None, "") else str(h).strip() for h in header]

        def idx_contains(*needles: str) -> int:
            needles_low = [n.lower() for n in needles if n]
            for i, h in enumerate(header):
                if not h:
                    continue
                hl = h.lower()
                if any(n in hl for n in needles_low):
                    return i
            raise KeyError(f"Missing column in city sheet {sheet}: contains {needles}")

        # Header cells are bilingual strings (Chinese + English).
        # Be careful: the sheet also contains "No. of family" which is numeric.
        def idx_family() -> int:
            candidates: list[tuple[int, str]] = []
            for i, h in enumerate(header):
                if not h:
                    continue
                hl = h.lower()
                if "family" in hl:
                    candidates.append((i, hl))
            # Prefer the bilingual " Family" column, not "No. of family".
            for i, hl in candidates:
                if "no." in hl and "family" in hl:
                    continue
                if hl.strip().endswith("family") or " family" in hl:
                    return i
            raise KeyError(f"Missing Family-like column in city sheet {sheet}")

        i_family = idx_family()
        i_sci = idx_contains("scientific name")
        i_native = idx_contains("native/non-native", "native / non-native", "native")

        for r in ws.iter_rows(min_row=3, values_only=True):
            sci = _norm(r[i_sci])
            fam = _norm(r[i_family]) or "Unknown"
            if not sci and fam == "Unknown":
                continue
            rows.append(
                {
                    "district": _norm(sheet),
                    "scientificName": sci,
                    "family": fam,
                    "nativeStatus": _norm(r[i_native]),
                }
            )

    return rows


def summarize_campus(rows: list[CampusRow]) -> dict[str, Any]:
    by_loc: dict[str, list[CampusRow]] = defaultdict(list)
    for r in rows:
        by_loc[r.locality].append(r)

    campuses: list[dict[str, Any]] = []
    for loc, rs in sorted(by_loc.items(), key=lambda kv: kv[0]):
        district = next((x.district for x in rs if x.district), "Unknown")
        lon = next((x.lon for x in rs if x.lon is not None), None)
        lat = next((x.lat for x in rs if x.lat is not None), None)

        species = {x.sci_name for x in rs if x.sci_name}
        families = {x.family for x in rs if x.family}
        growth = Counter(x.growth_form or "Unknown" for x in rs)

        native_flags = [_is_nonnative(x.nativeness) for x in rs]
        known = [x for x in native_flags if x is not None]
        nonnative_ratio = (sum(1 for x in known if x) / len(known)) if known else None

        top_families = Counter(x.family for x in rs if x.family).most_common(10)
        campuses.append(
            {
                "locality": loc,
                "district": district,
                "longitude": lon,
                "latitude": lat,
                "speciesCount": len(species),
                "familyCount": len(families),
                "nonnativeRatio": nonnative_ratio,
                "growthFormCounts": dict(growth),
                "topFamilies": [{"family": f, "count": c} for f, c in top_families],
            }
        )

    overall_species = len({r.sci_name for r in rows if r.sci_name})
    overall_families = len({r.family for r in rows if r.family})
    return {
        "kpi": {
            "campusCount": len(campuses),
            "speciesCount": overall_species,
            "familyCount": overall_families,
        },
        "campuses": campuses,
    }


def summarize_city(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_dist: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_dist[r["district"]].append(r)

    districts: list[dict[str, Any]] = []
    for dist, rs in sorted(by_dist.items(), key=lambda kv: kv[0]):
        species = {x["scientificName"] for x in rs if x["scientificName"]}
        families = {x["family"] for x in rs if x["family"]}
        flags = [_is_nonnative(x.get("nativeStatus")) for x in rs]
        known = [x for x in flags if x is not None]
        nonnative_ratio = (sum(1 for x in known if x) / len(known)) if known else None
        top_families = Counter(x["family"] for x in rs if x["family"]).most_common(10)
        districts.append(
            {
                "district": dist,
                "speciesCount": len(species),
                "familyCount": len(families),
                "nonnativeRatio": nonnative_ratio,
                "topFamilies": [{"family": f, "count": c} for f, c in top_families],
            }
        )

    overall_species = len({r["scientificName"] for r in rows if r["scientificName"]})
    overall_families = len({r["family"] for r in rows if r["family"]})
    return {
        "kpi": {
            "districtCount": len(districts),
            "speciesCount": overall_species,
            "familyCount": overall_families,
        },
        "districts": districts,
    }


def compute_overlap(
    campus_rows: list[CampusRow], city_rows: list[dict[str, Any]]
) -> dict[str, Any]:
    campus_fams: dict[str, set[str]] = defaultdict(set)
    campus_dist: dict[str, str] = {}
    for r in campus_rows:
        campus_fams[r.locality].add(r.family or "Unknown")
        if r.locality not in campus_dist and r.district:
            campus_dist[r.locality] = r.district

    city_fams: dict[str, set[str]] = defaultdict(set)
    for r in city_rows:
        city_fams[r["district"]].add(r["family"] or "Unknown")

    items: list[dict[str, Any]] = []
    for campus, fams in sorted(campus_fams.items(), key=lambda kv: kv[0]):
        dist = campus_dist.get(campus, "Unknown")
        dist_f = city_fams.get(dist, set())
        shared = fams & dist_f
        union = fams | dist_f
        jaccard = (len(shared) / len(union)) if union else None
        items.append(
            {
                "campus": campus,
                "district": dist,
                "campusFamilyCount": len(fams),
                "districtFamilyCount": len(dist_f),
                "sharedFamilyCount": len(shared),
                "jaccard": jaccard,
            }
        )

    return {"items": items}


def build_sankey(campus_rows: list[CampusRow]) -> dict[str, Any]:
    by_campus: dict[str, list[CampusRow]] = defaultdict(list)
    for r in campus_rows:
        by_campus[r.locality].append(r)

    nodes: dict[str, dict[str, Any]] = {}
    links: list[dict[str, Any]] = []

    def node(name: str, kind: str) -> None:
        if name not in nodes:
            nodes[name] = {"name": name, "kind": kind}

    for campus, rs in by_campus.items():
        dist = next((x.district for x in rs if x.district), "Unknown")
        node(dist, "district")
        node(campus, "campus")
        links.append({"source": dist, "target": campus, "value": len({x.sci_name for x in rs if x.sci_name})})

        fam_counts = Counter(x.family for x in rs if x.family).most_common(8)
        for fam, c in fam_counts:
            node(fam, "family")
            links.append({"source": campus, "target": fam, "value": c})

    return {"nodes": list(nodes.values()), "links": links}


def build_treemap(campus_rows: list[CampusRow], city_rows: list[dict[str, Any]]) -> dict[str, Any]:
    campus_counts = Counter(r.family for r in campus_rows if r.family)
    city_counts = Counter(r["family"] for r in city_rows if r.get("family"))

    def to_children(counter: Counter[str], n: int = 80) -> list[dict[str, Any]]:
        out = []
        for fam, c in counter.most_common(n):
            out.append({"name": fam, "value": int(c)})
        return out

    return {
        "campus": {"name": "Campus", "children": to_children(campus_counts)},
        "city": {"name": "Shanghai", "children": to_children(city_counts)},
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    campus_rows = load_campus_rows()
    city_rows = load_city_rows()

    campus_summary = summarize_campus(campus_rows)
    city_summary = summarize_city(city_rows)
    overlap = compute_overlap(campus_rows, city_rows)
    sankey = build_sankey(campus_rows)
    treemap = build_treemap(campus_rows, city_rows)

    (DATA_DIR / "campus_summary.json").write_text(
        json.dumps(campus_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "city_district_summary.json").write_text(
        json.dumps(city_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "overlap_family.json").write_text(
        json.dumps(overlap, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "sankey_district_campus_family.json").write_text(
        json.dumps(sankey, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "treemap_families.json").write_text(
        json.dumps(treemap, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print("Wrote JSON files to", DATA_DIR)


if __name__ == "__main__":
    main()

