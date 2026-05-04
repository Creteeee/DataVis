# -*- coding: utf-8 -*-
"""
Rebuild dashboard JSON assets and workflow-trace Excel files.

Outputs:
- data/campus_summary.json
- data/city_district_summary.json
- data/overlap_taxa.json
- data/rq2_taxa_native_nonnative.json
- data/overlap_family.json (compat alias from overlap_taxa.family)
- data/sankey_district_campus_family.json
- data/treemap_families.json
- workflow/Jaccard_Workflow_Species.xlsx
- workflow/Jaccard_Workflow_Genus.xlsx
- workflow/Jaccard_Workflow_Family.xlsx
- workflow/NativeNonnative_Stats_Species.xlsx
- workflow/NativeNonnative_Stats_Genus.xlsx
- workflow/NativeNonnative_Stats_Family.xlsx
"""

from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook


HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"
WORKFLOW_DIR = HERE / "workflow"

CAMPUS_XLSX = HERE.parent / "Plants_Shanghai_Translated.xlsx"
CITY_XLSX = HERE.parent / "上海野生和逸生植物_按地区拆分.xlsx"

DISTRICT_MAP = {
    "不详": "Unknown",
    "嘉定": "Jiading District",
    "崇明": "Chongming District",
    "宝山": "Baoshan District",
    "徐汇": "Xuhui District",
    "松江": "Songjiang District",
    "杨浦": "Yangpu District",
    "浦东": "Pudong New Area",
    "虹口": "Hongkou District",
    "长宁": "Changning District",
    "闵行": "Minhang District",
    "青浦": "Qingpu District",
    "静安": "Jing'an District",
    "黄浦": "Huangpu District",
    "金山": "Jinshan District",
    "奉贤": "Fengxian District",
    "普陀": "Putuo District",
}


def _norm(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _as_float(x: Any) -> float | None:
    if x is None:
        return None
    try:
        val = float(x)
        if math.isnan(val) or math.isinf(val):
            return None
        return val
    except Exception:
        return None


def _is_nonnative(value: Any) -> bool | None:
    s = _norm(value)
    if not s:
        return None
    low = s.lower()
    if "non" in low and "native" in low:
        return True
    if low == "nonnative":
        return True
    if low == "native":
        return False
    if any(k in s for k in ("非", "外来", "归化", "逸生")):
        return True
    if any(k in s for k in ("原生", "乡土")):
        return False
    return None


def _district_en(raw: str) -> str:
    s = _norm(raw)
    if not s:
        return "Unknown"
    if s in DISTRICT_MAP:
        return DISTRICT_MAP[s]
    # already English in campus file
    return s


def _species_key(scientific_name: str) -> str:
    """
    Canonical species token for cross-dataset matching.
    Keep the first two Latin-like tokens (genus + epithet), ignore authorship.
    """
    s = _norm(scientific_name)
    if not s:
        return ""
    parts = []
    for tok in s.replace("(", " ").replace(")", " ").replace(",", " ").split():
        t = tok.strip()
        if not t:
            continue
        # Latin-like token: letters and optional hyphen.
        if all(ch.isalpha() or ch == "-" for ch in t):
            parts.append(t.lower())
        if len(parts) >= 2:
            break
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}"
    return s.lower()


@dataclass(frozen=True)
class CampusRow:
    locality: str
    district: str
    lon: float | None
    lat: float | None
    species: str
    genus: str
    family: str
    growth_form: str
    nativeness: str


def _find_col(header: list[str | None], *, contains: str, excludes: tuple[str, ...] = ()) -> int:
    c = contains.lower()
    ex = tuple(e.lower() for e in excludes)
    for i, h in enumerate(header):
        if not h:
            continue
        hl = h.lower()
        if c in hl and not any(e in hl for e in ex):
            return i
    raise KeyError(f"Cannot find column containing {contains!r}")


def load_campus_rows() -> list[CampusRow]:
    wb = openpyxl.load_workbook(CAMPUS_XLSX, read_only=True, data_only=True)
    out: list[CampusRow] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        idx = {str(h).strip(): i for i, h in enumerate(header) if h not in (None, "")}

        i_locality = idx["Locality"]
        i_district = idx["District"]
        i_lon = idx["Longitude"]
        i_lat = idx["Latitude"]
        i_species = idx["ScientificName"]
        i_genus = idx["Genus"]
        i_family = idx["Family"]
        i_growth = idx["GrowthForm"]
        i_native = idx["Plant_NativenessStatus"]

        for r in ws.iter_rows(min_row=2, values_only=True):
            loc = _norm(r[i_locality])
            if not loc:
                continue
            out.append(
                CampusRow(
                    locality=loc,
                    district=_district_en(_norm(r[i_district]) or _norm(sheet)),
                    lon=_as_float(r[i_lon]),
                    lat=_as_float(r[i_lat]),
                    species=_norm(r[i_species]),
                    genus=_norm(r[i_genus]) or "Unknown",
                    family=_norm(r[i_family]) or "Unknown",
                    growth_form=_norm(r[i_growth]) or "Unknown",
                    nativeness=_norm(r[i_native]) or "Unknown",
                )
            )
    return out


def load_city_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(CITY_XLSX, read_only=True, data_only=True)
    out: list[dict[str, Any]] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_raw = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        header = [None if h in (None, "") else str(h).strip() for h in header_raw]

        i_species = _find_col(header, contains="scientific name")
        i_genus = _find_col(header, contains="genus", excludes=("chinese",))
        i_family = _find_col(header, contains="family", excludes=("chinese", "no. of family"))
        i_native = _find_col(header, contains="native/non-native")

        district_en = _district_en(_norm(sheet))
        for r in ws.iter_rows(min_row=3, values_only=True):
            species = _norm(r[i_species])
            genus = _norm(r[i_genus]) or "Unknown"
            family = _norm(r[i_family]) or "Unknown"
            if not species and genus == "Unknown" and family == "Unknown":
                continue
            out.append(
                {
                    "district": district_en,
                    "scientificName": species,
                    "genus": genus,
                    "family": family,
                    "nativeStatus": _norm(r[i_native]),
                }
            )
    return out


def summarize_campus(rows: list[CampusRow]) -> dict[str, Any]:
    by_loc: dict[str, list[CampusRow]] = defaultdict(list)
    for r in rows:
        by_loc[r.locality].append(r)

    campuses: list[dict[str, Any]] = []
    for loc, rs in sorted(by_loc.items(), key=lambda kv: kv[0]):
        species = {x.species for x in rs if x.species}
        genus = {x.genus for x in rs if x.genus}
        family = {x.family for x in rs if x.family}
        district = next((x.district for x in rs if x.district), "Unknown")
        lon = next((x.lon for x in rs if x.lon is not None), None)
        lat = next((x.lat for x in rs if x.lat is not None), None)
        growth = Counter(x.growth_form for x in rs)
        flags = [_is_nonnative(x.nativeness) for x in rs]
        known = [x for x in flags if x is not None]
        nonnative_ratio = (sum(1 for x in known if x) / len(known)) if known else None
        top_families = Counter(x.family for x in rs if x.family).most_common(10)

        campuses.append(
            {
                "locality": loc,
                "district": district,
                "longitude": lon,
                "latitude": lat,
                "speciesCount": len(species),
                "genusCount": len(genus),
                "familyCount": len(family),
                "nonnativeRatio": nonnative_ratio,
                "growthFormCounts": dict(growth),
                "topFamilies": [{"family": f, "count": c} for f, c in top_families],
            }
        )

    return {
        "kpi": {
            "campusCount": len(campuses),
            "speciesCount": len({r.species for r in rows if r.species}),
            "genusCount": len({r.genus for r in rows if r.genus}),
            "familyCount": len({r.family for r in rows if r.family}),
        },
        "campuses": campuses,
    }


def summarize_city(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_dist: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_dist[r["district"]].append(r)

    districts: list[dict[str, Any]] = []
    for dist, rs in sorted(by_dist.items(), key=lambda kv: kv[0]):
        species = {x["scientificName"] for x in rs if x["scientificName"]}
        genus = {x["genus"] for x in rs if x["genus"]}
        family = {x["family"] for x in rs if x["family"]}
        flags = [_is_nonnative(x.get("nativeStatus")) for x in rs]
        known = [x for x in flags if x is not None]
        nonnative_ratio = (sum(1 for x in known if x) / len(known)) if known else None
        top_families = Counter(x["family"] for x in rs if x["family"]).most_common(10)
        districts.append(
            {
                "district": dist,
                "speciesCount": len(species),
                "genusCount": len(genus),
                "familyCount": len(family),
                "nonnativeRatio": nonnative_ratio,
                "topFamilies": [{"family": f, "count": c} for f, c in top_families],
            }
        )

    return {
        "kpi": {
            "districtCount": len(districts),
            "speciesCount": len({r["scientificName"] for r in rows if r["scientificName"]}),
            "genusCount": len({r["genus"] for r in rows if r["genus"]}),
            "familyCount": len({r["family"] for r in rows if r["family"]}),
        },
        "districts": districts,
    }


def build_overlap_taxa(campus_rows: list[CampusRow], city_rows: list[dict[str, Any]]) -> dict[str, Any]:
    city_sets = {
        "species": defaultdict(set),
        "genus": defaultdict(set),
        "family": defaultdict(set),
    }
    for r in city_rows:
        city_sets["species"][r["district"]].add(_species_key(r["scientificName"]))
        city_sets["genus"][r["district"]].add(r["genus"])
        city_sets["family"][r["district"]].add(r["family"])

    by_campus: dict[str, list[CampusRow]] = defaultdict(list)
    for r in campus_rows:
        by_campus[r.locality].append(r)

    result: dict[str, Any] = {}
    for level in ("species", "genus", "family"):
        items: list[dict[str, Any]] = []
        for campus, rows in sorted(by_campus.items(), key=lambda kv: kv[0]):
            district = rows[0].district
            campus_set = {
                (
                    _species_key(x.species)
                    if level == "species"
                    else x.genus
                    if level == "genus"
                    else x.family
                )
                for x in rows
                if (
                    _species_key(x.species)
                    if level == "species"
                    else x.genus
                    if level == "genus"
                    else x.family
                )
            }
            district_set = city_sets[level].get(district, set())
            shared = campus_set & district_set
            union = campus_set | district_set
            items.append(
                {
                    "campus": campus,
                    "district": district,
                    "campusCount": len(campus_set),
                    "districtCount": len(district_set),
                    "sharedCount": len(shared),
                    "unionCount": len(union),
                    "jaccard": (len(shared) / len(union)) if union else None,
                }
            )
        result[level] = {"items": items}
    return result


# 桑基图左列行政区自上而下顺序（与 DISTRICT_MAP 英文一致）
SANKEY_DISTRICT_ORDER: tuple[str, ...] = (
    "Yangpu District",
    "Baoshan District",
    "Minhang District",
    "Fengxian District",
)


def _district_sankey_rank(name: str) -> int:
    try:
        return SANKEY_DISTRICT_ORDER.index(name)
    except ValueError:
        return 999


def build_sankey(campus_rows: list[CampusRow]) -> dict[str, Any]:
    """
    区→校与校→科均用去重物种数；仅 Top8 科；不生成「其余科」。
    区→校 value = 该校 Top8 科下去重物种数之和（与右侧流出一致）。
    nodes 顺序供 ECharts 同列垂直排列：区按 Yangpu→Baoshan→Minhang→Fengxian，校按区顺序，
    科按全图汇入总流量降序。
    """
    by_campus: dict[str, list[CampusRow]] = defaultdict(list)
    for r in campus_rows:
        by_campus[r.locality].append(r)

    node_meta: dict[str, dict[str, Any]] = {}
    links: list[dict[str, Any]] = []
    all_campuses: set[str] = set()
    districts_seen: set[str] = set()

    def ensure_node(name: str, kind: str) -> None:
        if name not in node_meta:
            node_meta[name] = {"name": name, "kind": kind}

    for campus, rs in by_campus.items():
        district = rs[0].district

        by_family: dict[str, set[str]] = defaultdict(set)
        for x in rs:
            if not x.species or not x.family:
                continue
            by_family[x.family].add(_species_key(x.species))
        top8 = sorted(by_family.items(), key=lambda kv: (-len(kv[1]), kv[0]))[:8]
        out_sum = sum(len(s) for _, s in top8)
        if out_sum == 0:
            continue
        districts_seen.add(district)
        all_campuses.add(campus)
        ensure_node(district, "district")
        ensure_node(campus, "campus")
        links.append({"source": district, "target": campus, "value": out_sum})
        for fam, sset in top8:
            if not sset:
                continue
            ensure_node(fam, "family")
            links.append({"source": campus, "target": fam, "value": len(sset)})

    # 科：全图总流量（所有校园→该科 value 之和），降序
    family_inflow: dict[str, int] = defaultdict(int)
    for l in links:
        if l["source"] in all_campuses:
            family_inflow[l["target"]] += int(l["value"])

    families_sorted = sorted(family_inflow.keys(), key=lambda f: (-family_inflow[f], f))

    # 区：按指定顺序，仅出现过的
    districts_sorted = sorted(
        districts_seen,
        key=lambda d: (_district_sankey_rank(d), d),
    )
    # 校：按区顺序，区内校名排序
    campus_by_district: dict[str, list[str]] = defaultdict(list)
    for c in by_campus:
        d = by_campus[c][0].district
        campus_by_district[d].append(c)
    for d in campus_by_district:
        campus_by_district[d].sort()

    ordered_nodes: list[dict[str, Any]] = []
    for d in districts_sorted:
        ordered_nodes.append(node_meta[d])
    for d in districts_sorted:
        for c in campus_by_district.get(d, []):
            if c in node_meta:
                ordered_nodes.append(node_meta[c])
    for fam in families_sorted:
        ordered_nodes.append(node_meta[fam])

    return {"nodes": ordered_nodes, "links": links}


def build_treemap(campus_rows: list[CampusRow], city_rows: list[dict[str, Any]]) -> dict[str, Any]:
    campus_counts = Counter(r.family for r in campus_rows if r.family)
    city_counts = Counter(r["family"] for r in city_rows if r.get("family"))

    def to_children(counter: Counter[str], n: int = 80) -> list[dict[str, Any]]:
        return [{"name": k, "value": int(v)} for k, v in counter.most_common(n)]

    return {
        "campus": {"name": "Campus", "children": to_children(campus_counts)},
        "city": {"name": "Shanghai", "children": to_children(city_counts)},
    }


def _taxon_value_from_campus(row: CampusRow, level: str) -> str:
    if level == "species":
        return _species_key(row.species)
    if level == "genus":
        return row.genus
    return row.family


def _taxon_value_from_city(row: dict[str, Any], level: str) -> str:
    if level == "species":
        return _species_key(row["scientificName"])
    if level == "genus":
        return row["genus"]
    return row["family"]


def _final_status(flags: list[bool | None]) -> str:
    known = [x for x in flags if x is not None]
    if not known:
        return "unknown"
    if any(x is True for x in known):
        return "nonnative"
    return "native"


def build_rq2_taxa_stats(campus_rows: list[CampusRow], city_rows: list[dict[str, Any]]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for level in ("species", "genus", "family"):
        # district taxon status
        district_taxon_flags: dict[str, dict[str, list[bool | None]]] = defaultdict(lambda: defaultdict(list))
        for r in city_rows:
            district = r["district"]
            taxon = _taxon_value_from_city(r, level)
            if not taxon:
                continue
            district_taxon_flags[district][taxon].append(_is_nonnative(r.get("nativeStatus")))

        district_items: list[dict[str, Any]] = []
        for district, taxa in sorted(district_taxon_flags.items(), key=lambda kv: kv[0]):
            native = nonnative = unknown = 0
            for flags in taxa.values():
                status = _final_status(flags)
                if status == "native":
                    native += 1
                elif status == "nonnative":
                    nonnative += 1
                else:
                    unknown += 1
            known_total = native + nonnative
            district_items.append(
                {
                    "district": district,
                    "nativeCount": native,
                    "nonnativeCount": nonnative,
                    "unknownCount": unknown,
                    "knownTotal": known_total,
                    "nonnativeRatio": (nonnative / known_total) if known_total else None,
                }
            )

        # campus taxon status
        campus_taxon_flags: dict[str, dict[str, list[bool | None]]] = defaultdict(lambda: defaultdict(list))
        campus_district: dict[str, str] = {}
        for r in campus_rows:
            campus = r.locality
            taxon = _taxon_value_from_campus(r, level)
            if not taxon:
                continue
            campus_taxon_flags[campus][taxon].append(_is_nonnative(r.nativeness))
            if campus not in campus_district:
                campus_district[campus] = r.district

        campus_items: list[dict[str, Any]] = []
        for campus, taxa in sorted(campus_taxon_flags.items(), key=lambda kv: kv[0]):
            native = nonnative = unknown = 0
            for flags in taxa.values():
                status = _final_status(flags)
                if status == "native":
                    native += 1
                elif status == "nonnative":
                    nonnative += 1
                else:
                    unknown += 1
            known_total = native + nonnative
            campus_items.append(
                {
                    "locality": campus,
                    "district": campus_district.get(campus, "Unknown"),
                    "nativeCount": native,
                    "nonnativeCount": nonnative,
                    "unknownCount": unknown,
                    "knownTotal": known_total,
                    "nonnativeRatio": (nonnative / known_total) if known_total else None,
                }
            )

        out[level] = {"districts": district_items, "campuses": campus_items}
    return out


def write_overlap_workbook(path: Path, level: str, items: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{level.title()} Jaccard"
    ws.append(
        [
            "Campus",
            "District",
            f"Campus_{level.title()}_Count",
            f"District_{level.title()}_Count",
            f"Shared_{level.title()}_Count",
            "Union_Count",
            "Jaccard",
        ]
    )
    for it in items:
        ws.append(
            [
                it["campus"],
                it["district"],
                it["campusCount"],
                it["districtCount"],
                it["sharedCount"],
                it["unionCount"],
                None if it["jaccard"] is None else round(it["jaccard"], 6),
            ]
        )
    wb.save(path)


def write_rq2_workbook(path: Path, level: str, rq2_level: dict[str, Any]) -> None:
    wb = Workbook()
    ws_d = wb.active
    ws_d.title = "District"
    ws_d.append(
        [
            "District",
            f"Native_{level.title()}_Count",
            f"Nonnative_{level.title()}_Count",
            f"Unknown_{level.title()}_Count",
            "Known_Total",
            "Nonnative_Ratio",
        ]
    )
    for r in rq2_level["districts"]:
        ws_d.append(
            [
                r["district"],
                r["nativeCount"],
                r["nonnativeCount"],
                r["unknownCount"],
                r["knownTotal"],
                None if r["nonnativeRatio"] is None else round(r["nonnativeRatio"], 6),
            ]
        )

    ws_c = wb.create_sheet("Campus")
    ws_c.append(
        [
            "Campus",
            "District",
            f"Native_{level.title()}_Count",
            f"Nonnative_{level.title()}_Count",
            f"Unknown_{level.title()}_Count",
            "Known_Total",
            "Nonnative_Ratio",
        ]
    )
    for r in rq2_level["campuses"]:
        ws_c.append(
            [
                r["locality"],
                r["district"],
                r["nativeCount"],
                r["nonnativeCount"],
                r["unknownCount"],
                r["knownTotal"],
                None if r["nonnativeRatio"] is None else round(r["nonnativeRatio"], 6),
            ]
        )
    wb.save(path)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    WORKFLOW_DIR.mkdir(parents=True, exist_ok=True)

    campus_rows = load_campus_rows()
    city_rows = load_city_rows()

    campus_summary = summarize_campus(campus_rows)
    city_summary = summarize_city(city_rows)
    overlap_taxa = build_overlap_taxa(campus_rows, city_rows)
    rq2_taxa = build_rq2_taxa_stats(campus_rows, city_rows)
    sankey = build_sankey(campus_rows)
    treemap = build_treemap(campus_rows, city_rows)

    (DATA_DIR / "campus_summary.json").write_text(
        json.dumps(campus_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "city_district_summary.json").write_text(
        json.dumps(city_summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "overlap_taxa.json").write_text(
        json.dumps(overlap_taxa, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "rq2_taxa_native_nonnative.json").write_text(
        json.dumps(rq2_taxa, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "overlap_family.json").write_text(
        json.dumps(overlap_taxa["family"], ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "sankey_district_campus_family.json").write_text(
        json.dumps(sankey, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "treemap_families.json").write_text(
        json.dumps(treemap, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    write_overlap_workbook(
        WORKFLOW_DIR / "Jaccard_Workflow_Species.xlsx",
        "species",
        overlap_taxa["species"]["items"],
    )
    write_overlap_workbook(
        WORKFLOW_DIR / "Jaccard_Workflow_Genus.xlsx",
        "genus",
        overlap_taxa["genus"]["items"],
    )
    write_overlap_workbook(
        WORKFLOW_DIR / "Jaccard_Workflow_Family.xlsx",
        "family",
        overlap_taxa["family"]["items"],
    )
    write_rq2_workbook(
        WORKFLOW_DIR / "NativeNonnative_Stats_Species.xlsx",
        "species",
        rq2_taxa["species"],
    )
    write_rq2_workbook(
        WORKFLOW_DIR / "NativeNonnative_Stats_Genus.xlsx",
        "genus",
        rq2_taxa["genus"],
    )
    write_rq2_workbook(
        WORKFLOW_DIR / "NativeNonnative_Stats_Family.xlsx",
        "family",
        rq2_taxa["family"],
    )

    print(f"Done. JSON -> {DATA_DIR}")
    print(f"Done. Excel workflow -> {WORKFLOW_DIR}")


if __name__ == "__main__":
    main()

